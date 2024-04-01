import pandas as pd

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from copy import copy
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.styles import Border, Side
from openpyxl.drawing.xdr import XDRPositiveSize2D, XDRPoint2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor

from io import BytesIO

import streamlit as st


# Cabeçalhos para a primeira linha de cada aba
cabecalhos = [
    'REF.', 'IMAGEM', 'DESCRIÇÃO', 'custo sergio', 'custo paulo', 'Grupo', 'Subgrupo', 
    'Codigo de barras', 'DUN 14', 'CLEINTE', 'NCM', 'II', 'IPI', 'PIS', 'COFINS', 
    'PACKAGE', 'MATERIAL', 'PREÇO', 'UNIT', 'PCS/CT', 'CX', 'L', 'W', 'H', 'CBM', 
    'PESO/UNITARIO(G)', 'G.W. (KG)', 'N.W. (KG)', 'T*GW', 'T*NW', 'T*QUANT', 'T*CBM', 
    'T*VALOR', 'OBS', ' ', 'VALOR FOB',	'DECLARADO', 'POR FORA', 'CAMBIO OFICIAL', 	
    'CAMBIO FORA', 	'IMPORTAÇÃO', 'TOTAL OFICIAL', 'PC OFICIAL', 'TOTAL REAL',	
    'PC REAL', 	'imposto saida', 'banana', 	'custo sergio', 'custo paulo'

]

# Definir o estilo de borda
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Criar uma função auxiliar para aplicar a formatação
def aplicar_formato_celula(celula_destino, celula_origem):
    # Copiar a formatação numérica e o estilo
    celula_destino.number_format = celula_origem.number_format
    
    # Copiar a fonte, alinhamento, cor de fundo e borda
    celula_destino.font = copy(celula_origem.font)
    celula_destino.alignment = copy(celula_origem.alignment)
    celula_destino.fill = copy(celula_origem.fill)
    celula_destino.border = copy(celula_origem.border)



def processar_planilha(uploaded_file):
    wb_original = load_workbook(uploaded_file, data_only=True)
    sheet_original = wb_original.active
    
    # Identificar subgrupos e associar imagens
    subgrupos = {}
    imagens_subgrupo = {}
    dados_subgrupo = {}  # Novo dicionário para armazenar os dados associados a cada subgrupo


         ###################       
    for img in sheet_original._images:
        # Achar a célula referente à imagem
        ref_row = img.anchor._from.row
        
        # Verificar se ref_row é maior que 1 para processar a partir da segunda linha
        if ref_row >= 1:
            subgrupo = sheet_original.cell(row=ref_row+1, column=7).value
            
            # Substituir subgrupo com "?" por "TESTE"
            if '?' in subgrupo:
                subgrupo = "unknown"
            
            if subgrupo:
                # Coletar dados da linha
                dados_linha = [sheet_original.cell(row=ref_row+1, column=col).value for col in range(1, sheet_original.max_column + 1)]
                
                if subgrupo not in subgrupos:
                    print("Subgrupo: ", subgrupo)
                    subgrupos[subgrupo] = []
                    imagens_subgrupo[subgrupo] = []
                    dados_subgrupo[subgrupo] = []  # Inicializar a lista de dados para este subgrupo
                
                imagens_subgrupo[subgrupo].append(img)
                dados_subgrupo[subgrupo].append(dados_linha)  # Adicionar os dados da linha ao subgrupo correspondente

    # Seu código de processamento vem aqui
    # Nota: Devido à complexidade, esta é uma representação simplificada do seu código

    # Para o propósito de simplificação, vamos simular a criação de um novo workbook
    wb_novo = Workbook()
    ws_novo = wb_novo.active
    ws_novo.title = "DadosProcessados"

    # Configurações de tamanho para as imagens
    altura_cm = 2.8
    largura_cm = 3.2
    pixels_por_cm = 35  # Valor aproximado para conversão
    altura_px = int(altura_cm * pixels_por_cm)
    largura_px = int(largura_cm * pixels_por_cm)

    # Adicionar cabeçalhos e imagens nas abas correspondentes aos subgrupos
    for subgrupo, imgs in imagens_subgrupo.items():
        ws_subgrupo = wb_novo.create_sheet(title=subgrupo)
        
        # Escrever os cabeçalhos na primeira linha
        for col_index, cabecalho in enumerate(cabecalhos, start=1):
            ws_subgrupo.cell(row=1, column=col_index, value=cabecalho)

        # Aplicar um filtro em todas as colunas com cabeçalho
        ws_subgrupo.auto_filter.ref = f"A1:{get_column_letter(len(cabecalhos))}{len(imgs)+1}"
        
        for index, (img, dados_linha) in enumerate(zip(imgs, dados_subgrupo[subgrupo]), start=2):
            # Adicionar dados da linha na aba
            for col_index, valor in enumerate(dados_linha, start=1):
                # Atribui CELL a nova celula cm os 3 valores: linha/coluna/valor
                cell = ws_subgrupo.cell(row=index, column=col_index, value=valor)
                
                # Copiar a formatação numérica e o estilo 
                original_cell = sheet_original.cell(row=ref_row+1, column=col_index)

                aplicar_formato_celula (cell, original_cell)
            
            # Adicionar a imagem na célula referenciada
            img_copia = Image(img.ref)
            img_copia.width = largura_px
            img_copia.height = altura_px
            cell_ref = f'B{index}'  # Posicionar na coluna 'B' pois 'A' é para REF.
            ws_subgrupo.add_image(img_copia, cell_ref)
            
            # Ajustar altura da linha e largura da coluna para as imagens
            ws_subgrupo.row_dimensions[index].height = 90        
            ws_subgrupo.column_dimensions[get_column_letter(2)].width = 20

                # Após adicionar todos os dados e imagens na aba, definir bordas para todas as células
        for row in ws_subgrupo.iter_rows(min_row=1, max_col=ws_subgrupo.max_column, max_row=ws_subgrupo.max_row):
            for cell in row:
                cell.border = thin_border

    # Após adicionar todos os dados e imagens na aba
    for subgrupo, imgs in imagens_subgrupo.items():
        ws_subgrupo = wb_novo[subgrupo]

        # Iterar por todas as colunas com dados
        for col in ws_subgrupo.columns:
            max_length = 0
            column = col[0].column  # Pegar o número da coluna (1, 2, 3, ...)
            
            for cell in col:
                try:  # Necessário para evitar erro no caso de uma célula conter None
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ajustar a largura da coluna aqui, o fator 1.2 é um multiplicador para dar algum espaço extra
            ws_subgrupo.column_dimensions[get_column_letter(column)].width = max_length * 1.2
            if (ws_subgrupo.column_dimensions[get_column_letter(column)].width < 12):
                ws_subgrupo.column_dimensions[get_column_letter(column)].width = 13
            ws_subgrupo.column_dimensions[get_column_letter(2)].width = 20
        
    # Salvando o novo workbook em um BytesIO (em memória)
    output = BytesIO()
    wb_novo.save(output)
    output.seek(0)
    
    return output


# Streamlit App
def main():
    st.title('Upload e Processamento de Planilha XLSX com Streamlit')

    uploaded_file = st.file_uploader("Faça upload de sua planilha XLSX", type="xlsx")
    
    if uploaded_file is not None:
        processed_file = processar_planilha(uploaded_file)
        
        st.success('Planilha processada com sucesso!')
        st.download_button(label="Baixar Planilha Ajustada",
                           data=processed_file,
                           file_name="planilha_ajustada2.xlsx")

if __name__ == "__main__":
    main()